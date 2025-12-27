from user.models import User
from user.serializers import UserSerializers
from attendance.models import Attendance
from contact.models import Contact
from contact.serializers import ContactSerializers
from church.models import Church
from membership.models import Membership
from person.models import Person
from person.serializers import PersonSerializers
from role.models import Role
from capturemethod.models import CaptureMethod
from services.models import Services
from household.models import HouseHold
from household.serializers import HouseHoldSerializers
from attendance.serializers import attendanceSerializers
from django.shortcuts import render
from rest_framework import status
from rest_framework.reverse import reverse
from rest_framework import generics
from rest_framework.response import Response
from rest_framework import filters
from django_filters.rest_framework import DjangoFilterBackend
from urllib.parse import urlparse
from rest_framework.permissions import IsAuthenticated, IsAdminUser,AllowAny
from django_filters import AllValuesFilter, DateTimeFilter, NumberFilter
from rest_framework.exceptions import PermissionDenied
from django.http import HttpResponse
from role.util import requiredGroups
from user.permissions import IsInGroup
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from io import BytesIO


def addStyles(ws):
    bold_font = Font(bold=True)
    for row in ws.iter_rows():
        for cell in row:
           cell.alignment = Alignment(wrap_text=True)
           # Assign the bold_font object to the cell's font attribute
           cell.font = bold_font
def adjustWidth(ws):
        # Iterate over all columns to adjust their widths
        for col in ws.columns:
          max_length = 0
          column_letter = col[0].column_letter # Get the column letter (e.g., 'A', 'B')
          for cell in col:
            try:
            # Check the length of the cell's value
               if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
            except TypeError: # Handle cases where cell.value might be None
              pass
            # Calculate adjusted width (add a small buffer for visual spacing)
            adjusted_width = (max_length + 2)
            # Set the column width
            ws.column_dimensions[column_letter].width = adjusted_width

#this generic class will handle GET method to be used by the admin alone
class AttendanceList(generics.ListAPIView):
    queryset = Attendance.objects.all()
    serializer_class = attendanceSerializers
    permission_classes = [IsAuthenticated,IsInGroup]
    required_groups = requiredGroups(permission='view_attendance')
    name = 'attendance-report'

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    #you can filter by field names specified here keyword e.g url?name='church one'
    filterset_fields = ('comment','checkInTimestamp','checkOutTimestamp',
                        'personId__id','servicesId__id','captureMethodId__id') 

     #you can search using the "search" keyword
    search_fields = ('comment','checkInTimestamp','checkOutTimestamp',
                        'personId__id','servicesId__id','captureMethodId__id')

    #you can order using the "ordering" keyword
    ordering_fields = ('comment','checkInTimestamp','checkOutTimestamp',
                        'personId__id','servicesId__id','captureMethodId__id') 
    
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(queryset=self.get_queryset()) #still filter
        customQueryset = self.getCustomQueryset(queryset)
        try:
             # 3. Create an Excel workbook
              wb = Workbook()
              ws = wb.active
              ws.title = "Church Attendance"
              # Add headers (using serializer field names or custom names)
              headers =  [
                 'SN','FirstName','LastName', 'Service','CaptureMethod',
                 'CheckInTimestamp','CheckOutTimestamp','Remark',
              ]
              ws.append(headers)
              # Add data rows
              for item in customQueryset:
                   row_values = [item.get(header) for header in headers] # Or customize based on item keys
                   ws.append(row_values)
              #add styles
              addStyles(ws)
              adjustWidth(ws)
              # 4. Save the workbook to a BytesIO buffer
              excel_buffer = BytesIO()
              wb.save(excel_buffer)
              excel_buffer.seek(0) # Rewind the buffer to the beginning
              # 5. Create an HttpResponse with appropriate headers
              response = HttpResponse(
                 excel_buffer.getvalue(),
                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
              )
              response['Content-Disposition'] = 'attachment; filename="church_attendance.xlsx"'
              return response
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def getCustomQueryset(self, queryset):
         data = []
         count = 1
         for obj in queryset:
             capturemethod = CaptureMethod.objects.get(id=obj.captureMethodId.id)
             person = Person.objects.get(pk=obj.personId.id)
             service = Services.objects.get(id=obj.servicesId.id)
             checkout = None
             if obj.checkOutTimestamp is not None:
                checkout = obj.checkOutTimestamp.replace(tzinfo=None)

             newObj = {
                 'SN':count,'FirstName':person.firstName,
                 'LastName':person.lastName, 'Service': service.eventName,
                 'CaptureMethod': capturemethod.method,
                 'CheckInTimestamp': obj.checkInTimestamp.replace(tzinfo=None),
                 'CheckOutTimestamp': checkout,
                 'Remark': obj.comment,
             }
             data.append(newObj)
             count = count + 1
         return data
    


class HouseHoldList(generics.ListAPIView):
    queryset = HouseHold.objects.all()
    serializer_class = HouseHoldSerializers
    permission_classes = [IsAuthenticated,IsInGroup]
    required_groups = requiredGroups(permission='view_household')
    name = 'household-report'

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    #you can filter by field names specified here keyword e.g url?name='church one'
    filterset_fields = ('name','address','head','spouse','children','count') 

     #you can search using the "search" keyword
    search_fields = ('name','address','head','spouse','children','count')  

    #you can order using the "ordering" keyword
    ordering_fields = ('name','address','head','spouse','children','count') 

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(queryset=self.get_queryset()) #still filter
        customQueryset = self.getCustomQueryset(queryset)
        try:
             # 3. Create an Excel workbook
              wb = Workbook()
              ws = wb.active
              ws.title = "Church Families"
              # Add headers (using serializer field names or custom names)
              headers =  [
                 'SN','Name','Address', 'Count','Head',
                 'Spouse','Children',
              ]
              ws.append(headers)
              # Add data rows
              for item in customQueryset:
                   row_values = [item.get(header) for header in headers] # Or customize based on item keys
                   ws.append(row_values)
              #add styles
              addStyles(ws)
              adjustWidth(ws)
              # 4. Save the workbook to a BytesIO buffer
              excel_buffer = BytesIO()
              wb.save(excel_buffer)
              excel_buffer.seek(0) # Rewind the buffer to the beginning
              # 5. Create an HttpResponse with appropriate headers
              response = HttpResponse(
                 excel_buffer.getvalue(),
                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
              )
              response['Content-Disposition'] = 'attachment; filename="household_report.xlsx"'
              return response
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def getCustomQueryset(self, queryset):
         data = []
         count = 1
         for obj in queryset:
             newObj = {
                 'SN':count,'Name':obj.name,
                 'Address':obj.address, 'Count':obj.count,'Head':obj.head,
                 'Spouse':obj.spouse,'Children':obj.children,
             }
             data.append(newObj)
             count = count + 1
         return data
    


class UserList(generics.ListAPIView):
    queryset = User.objects.all()
    serializer_class = UserSerializers
    permission_classes = [AllowAny,]
    #required_groups = requiredGroups(permission='view_user')
    name = 'user-report'

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    #you can filter by field names specified here keyword e.g url?name='church one'
    filterset_fields = ('username','email','personId__id','roleId__id') 

     #you can search using the "search" keyword
    search_fields =  ('username','email','personId__id','roleId__id')  

    #you can order using the "ordering" keyword
    ordering_fields =  ('username','email','personId__id','roleId__id') 

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(queryset=self.get_queryset()) #still filter
        customQueryset = self.getCustomQueryset(queryset)
        try:
             # 3. Create an Excel workbook
              wb = Workbook()
              ws = wb.active
              ws.title = "Users Accounts"
              # Add headers (using serializer field names or custom names)
              headers =  [
                 'SN','FirstName','LastName', 'Username','Email',
                 'Role',
              ]
              ws.append(headers)
              # Add data rows
              for item in customQueryset:
                   row_values = [item.get(header) for header in headers] # Or customize based on item keys
                   ws.append(row_values)
              #add styles
              addStyles(ws)
              adjustWidth(ws)
              # 4. Save the workbook to a BytesIO buffer
              excel_buffer = BytesIO()
              wb.save(excel_buffer)
              excel_buffer.seek(0) # Rewind the buffer to the beginning
              # 5. Create an HttpResponse with appropriate headers
              response = HttpResponse(
                 excel_buffer.getvalue(),
                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
              )
              response['Content-Disposition'] = 'attachment; filename="user_report.xlsx"'
              return response
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def getCustomQueryset(self, queryset):
         data = []
         count = 1
         for obj in queryset:
             lastName = None
             FirstName = None
             role = None
             if(obj.personId.id is not None):
                 person = Person.objects.get(pk=obj.personId.id)
                 lastName = person.lastName
                 FirstName = person.firstName
            
             if(obj.roleId.id is not None):
                 role = Role.objects.get(id=obj.roleId.id).name

             newObj = {
                 'SN':count,'FirstName':FirstName,
                 'LastName':lastName, 'Username':obj.username,
                 'Email':obj.email, 'Role': role,
             }
             data.append(newObj)
             count = count + 1
         return data



class PersonList(generics.ListAPIView):
    queryset = Person.objects.all()
    serializer_class = PersonSerializers
    permission_classes = [AllowAny,]
    #required_groups = requiredGroups(permission='view_person')
    name = 'person-report'

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    #you can filter by field names specified here keyword e.g url?name='church one'
    filterset_fields = ('churchId__id','householdId__id','membershipId__id',
                        'firstName','lastName','middleName','dob','phone',
                        'email','entranceDate') 

     #you can search using the "search" keyword
    search_fields =  ('churchId__id','householdId__id','membershipId__id',
                        'firstName','lastName','middleName','dob','phone',
                        'email','entranceDate')  

    #you can order using the "ordering" keyword
    ordering_fields =  ('churchId__id','householdId__id','membershipId__id',
                        'firstName','lastName','middleName','dob','phone',
                        'email','entranceDate') 
    

    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(queryset=self.get_queryset()) #still filter
        customQueryset = self.getCustomQueryset(queryset)
        try:
             # 3. Create an Excel workbook
              wb = Workbook()
              ws = wb.active
              ws.title = "Church Members"
              # Add headers (using serializer field names or custom names)
              headers =  [
                 'SN','Church','Household','Membership Status',
                        'FirstName','LastName','MiddleName','DOB','Phone',
                        'Email','EntranceDate'
              ]
              ws.append(headers)
              # Add data rows
              for item in customQueryset:
                   row_values = [item.get(header) for header in headers] # Or customize based on item keys
                   ws.append(row_values)
              #add styles
              addStyles(ws)
              adjustWidth(ws)
              # 4. Save the workbook to a BytesIO buffer
              excel_buffer = BytesIO()
              wb.save(excel_buffer)
              excel_buffer.seek(0) # Rewind the buffer to the beginning
              # 5. Create an HttpResponse with appropriate headers
              response = HttpResponse(
                 excel_buffer.getvalue(),
                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
              )
              response['Content-Disposition'] = 'attachment; filename="members_report.xlsx"'
              return response
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def getCustomQueryset(self, queryset):
         data = []
         count = 1
         for obj in queryset:
             status = None
             if(obj.membershipId.id is not None):
                 status = Membership.objects.get(id=obj.membershipId.id).status
             church = Church.objects.get(id=obj.churchId.id)
             household = HouseHold.objects.get(id=obj.householdId.id)

             dob = None
             if obj.dob is not None:
                dob = obj.dob

             entranceDate = None
             if obj.entranceDate is not None:
                entranceDate = obj.entranceDate.replace(tzinfo=None)

             newObj = {
                 'SN':count,'Church':church.name,
                 'Household':household.name, 'Membership Status':status,
                 'FirstName':obj.firstName,
                 'LastName':obj.lastName, 'MiddleName':obj.middleName,
                 'DOB':dob, 'Phone':obj.phone,
                 'Email':obj.email, 'EntranceDate':entranceDate
             }
             data.append(newObj)
             count = count + 1
         return data
    


class ContactList(generics.ListAPIView):
    queryset = Contact.objects.all()
    serializer_class = ContactSerializers
    permission_classes = [IsAuthenticated,IsInGroup]
    required_groups = requiredGroups(permission='view_contact')
    name = 'contact-report'

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    #you can filter by field names specified here keyword e.g url?name='church one'
    filterset_fields = ('personId__id','address','householdId__id','email','phone','socialMedia','gender',
                        'ethnicity','state','occupation', 'country','marital_status') 

     #you can search using the "search" keyword
    search_fields = ('personId__id','address','householdId__id','email','phone','socialMedia','gender',
                        'ethnicity','state','occupation', 'country','marital_status')  

    #you can order using the "ordering" keyword
    ordering_fields = ('personId__id','address','householdId__id','email','phone','socialMedia','gender',
                        'ethnicity','state','occupation', 'country','marital_status')
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(queryset=self.get_queryset()) #still filter
        customQueryset = self.getCustomQueryset(queryset)
        try:
             # 3. Create an Excel workbook
              wb = Workbook()
              ws = wb.active
              ws.title = "Members Contacts"
              # Add headers (using serializer field names or custom names)
              headers =  [
                 'SN','FirstName','LastName','Address','Household','Email','Phone','SocialMedia','Gender',
                        'Ethnicity','State','Occupation', 'Country','Marital_Status'
              ]
              ws.append(headers)
              # Add data rows
              for item in customQueryset:
                   row_values = [item.get(header) for header in headers] # Or customize based on item keys
                   ws.append(row_values)
              #add styles
              addStyles(ws)
              adjustWidth(ws)
              # 4. Save the workbook to a BytesIO buffer
              excel_buffer = BytesIO()
              wb.save(excel_buffer)
              excel_buffer.seek(0) # Rewind the buffer to the beginning
              # 5. Create an HttpResponse with appropriate headers
              response = HttpResponse(
                 excel_buffer.getvalue(),
                 content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
              )
              response['Content-Disposition'] = 'attachment; filename="members_contacts.xlsx"'
              return response
        except Exception as e:
            return Response({"error": str(e)}, status=500)

    def getCustomQueryset(self, queryset):
         data = []
         count = 1
         for obj in queryset:
             person = Person.objects.get(id=obj.personId.id)
             household = HouseHold.objects.get(id=obj.householdId.id)

             newObj = {
                 'SN': count, 'FirstName': person.firstName,
                 'LastName': person.lastName, 'Phone': obj.phone,
                 'Email': obj.email, 'Address': obj.address,
                 'Household': household.name, 'SocialMedia': obj.socialMedia,
                 'Gender': obj.gender, 'Ethnicity': obj.ethnicity,
                 'State': obj.state, 'Occupation': obj.occupation,
                 'Country': obj.country, 'Marital_Status': obj.marital_status
             }
             data.append(newObj)
             count = count + 1
         return data